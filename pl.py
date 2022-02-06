import openpyxl
import math
from openpyxl.styles import Alignment, PatternFill
import collections
class Solution:
    def foo(self):
        file=str(input("File name: "))
        wb = openpyxl.load_workbook(file)
        sh1=str(input("Sheet 1: "))
        sh2=str((input("VLAN to work on: ")))
        if sh2 not in wb.sheetnames:
            wb.create_sheet(sh2)
        c=0 #counter
        
        ws1=wb[sh1]
        ws2=wb[sh2]
        
        map={0:0,1:1,2:5,3:6}
        d=collections.defaultdict(list)
        ls=dict()
        ws2.cell(row=1, column=1).value="SRC_IP"
        ws2.cell(row=1, column=2).value="SRC_Name"
        ws2.cell(row=1, column=3).value="DST_IP"
        ws2.cell(row=1, column=4).value="DST_Name"
        ws2.cell(row=1, column=5).value="PROTOCOL"
        ws2.cell(row=1, column=6).value="PORT"
        
        for rows in ws2.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
            for cell in rows:
                cell.fill = PatternFill(start_color='00CCFFFF', end_color='00CCFFFF', fill_type="solid")
        
        for i,val in enumerate(ws1.iter_rows(min_row=2, min_col=1, values_only=True)):
            if sh2==val[2]:
                if val[5] is None:
                    continue
                d[(val[0],val[1])].append((val[5],val[6] or "--Null--",val[11],val[10]))
                print(val)
            else:
                continue
    
        
        for (key,value) in d.items():
            if len(value)>=27:
                for i in range(c+2, c+2+math.ceil(len(value)/27)):
                    ws2.row_dimensions[i].height=400
                ws2.merge_cells(start_row=c+2, end_row=c+1+math.ceil(len(value)/27), start_column=3, end_column=3) #0+1+2=3 2--3
                ws2.merge_cells(start_row=c+2, end_row=c+1+math.ceil(len(value)/27), start_column=4, end_column=4)
                ws2.merge_cells(start_row=c+2, end_row=c+1+math.ceil(len(value)/27), start_column=5, end_column=5)
                ws2.merge_cells(start_row=c+2, end_row=c+1+math.ceil(len(value)/27), start_column=6, end_column=6)
             
            ws2.cell(row=c+2, column=1).value=key[0] 
            #print(c+2,math.ceil(len(value)/27)+c+1)
            ws2.cell(row=c+2, column=2).value=key[1]
            ws2.cell(row=c+2, column=3).value="\n".join([item[0] for item in value])
            ws2.cell(row=c+2, column=4).value="\n".join([item[1] for item in value])
            ws2.cell(row=c+2, column=5).value="\n".join([item[2] for item in value])
            ws2.cell(row=c+2, column=6).value="\n".join([item[3] for item in value])
            
            c+=math.ceil(len(value)/27) #c=0 --> 0+2=2 --> c=2
            
            
        ws2.alignment=Alignment(wrap_text=True)
        wb.save(file)
        print("Done!")
  
        

     
p=Solution()
p.foo()

import openpyxl
import collections
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.borders import Border, Side
class Solution(object):
    def foo(self):
        try:
            file1='C:/Users/[USERNAME]/data.xlsx'
            file1='C:/Users/[USERNAME]/data2.xlsx'
            wb1=openpyxl.load_workbook(file1)
            wb2=openpyxl.load_workbook(file2)
            ws1=wb1.worksheets[0]
            ws2=wb2.worksheets[0]
            ls=[]
            cls=[]
            """
            --------------------------------------------------------------------------------------
            | Creating a dictionary for port/protocol thing and write up the head of a new sheet.|
            --------------------------------------------------------------------------------------
            """
            d=collections.defaultdict(list)
            thin_border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            if ws2.cell(row=1, column=1).value is None:
                ws2.cell(row=1, column=1).value="Source IP addresses"
                ws2.cell(row=1, column=2).value="Name"
                ws2.cell(row=1, column=3).value="Destination IP addresses"
                ws2.cell(row=1, column=4).value="Name"
                ws2.cell(row=1, column=5).value="Protocol/Port"
                ws2.cell(row=1, column=6).value="Service or application"
                ws2.cell(row=1, column=7).value="Comments"
                ws2.cell(row=1, column=8).value="Call"
                ws2.cell(row=1, column=9).value="Proposed FW rule ID \[EY\]"
                
            """
            -------------------------------------------------------------------------------
            | All the data collected from the source table are saving in lists (seperate) |
            -------------------------------------------------------------------------------
            """

            sip,sname,dip,dname,port,pr=list(),list(),list(),list(),list(),list()
            for row in ws1.iter_rows(min_row=2):



                if row[0].value in sip:
                    pass
                else:
                    sip.append(row[0].value)
                if row[1].value in sname:
                    pass
                else:
                    sname.append(row[1].value)
                if row[5].value in dip:
                    pass
                else:
                    dip.append(row[5].value)
                if row[6].value in dname:
                    pass
                else:
                    dname.append(row[6].value)
                if row[10].value in port:
                    pass
                else:
                    port.append(row[10].value)
                if row[11].value in pr:
                    pass
                else:
                    pr.append(row[11].value)

                if row[10].value in cls:
                    continue
                d[row[11].value].append(row[10].value)
                cls.append(row[10].value)


            for key,value in d.items():

                if len(value)==1:
                    ls.append(key)
                    ls.append('/')
                    ls.append(value[0])
                if len(value)>1:
                    for i in range(len(value)):
                        ls.append(key)
                        ls.append('/')
                        ls.append(value[i])

            nls=[]

            for num,val in enumerate(ls):
                if (num)%3==0:
                    nls.append("\n")
                nls.append(val)

            for i in range(2,1000):
                if ws2.cell(row=i, column=1).value is None:
                    if i%2==0:
                        for col in range(1,10):
                            ws2.cell(row=i, column=col).border = thin_border
                            ws2.cell(row=i, column=col).fill=PatternFill(start_color='0099CCFF', end_color='0099CCFF', fill_type='solid')
                    else:
                        for col in range(1,10):
                            ws2.cell(row=i, column=col).fill=PatternFill(start_color='00CCFFCC', end_color='00CCFFCC', fill_type='solid')
                            ws2.cell(row=i, column=col).border = thin_border
                    ws2.cell(row=i, column=1).value="\n".join([i for i in sip if i!=None])
                    ws2.cell(row=i, column=2).value="\n".join([i for i in sname if i!=None])
                    ws2.cell(row=i, column=3).value="\n".join([i for i in dip if i!=None])
                    ws2.cell(row=i, column=4).value="\n".join([i for i in dname if i!=None])
                    ws2.cell(row=i, column=5).value="".join([str(i) for i in nls if i!=None])

                    break
            ws2.alignment=Alignment(wrap_text=True)
            wb2.save(file2)
            print("""MMMMMMWNKOxddollllllloodx0WMMMMMMMMMMMWX0xolc::::clox0NWMMMMMMMMMMMMMKdllloOWMMMMMMMNOollllokNMMMMMMM0olloollollllllllkN
                    MMMWXko:,''''''''''''''''dNMMMMMMMMMW0d:,'''''''''''',:dKWMMMMMMMMMMWk,''''dNMMMMMMNd,''''''cKMMMMMMWx,'''''''''''''''cK
                    MMNk:'''''',;cloooc;'''''dNMMMMMMMMXd;''''',coddol;''''';dXMMMMMMMMMWk,''''dNMMMMWKo,'''''''cKMMMMMMWOlcccccccc:;'''''cK
                    MXo,'''''cx0XWWMMWXl'''''dNMMMMMMMKl''''',o0NWMMMWKx;'''''lKWWMMMMMMWk,''''dNMMMW0c'.'''''''cKMMMMMMMWWNNNNNNNNNd,''''cK
                    Nd,'''',dXMMMMMMMMXl'''''dNMMMMMMXo''''';xNMMMMMMMMWO:'''''lOXMMMMMMWk,''''dNMMNk:''''''''''cKMMMMMMMMMMMMMMMMMWx,''''cK
                    0:'''''lXMMMMMMMMMXl'''''dNMMMMMMO;'''''oXMMMMMMMMMMWk,'''',:kWMMMMMWk,''''dNMXd,'''':o:''''cKMMMMMMMW0kkkkkkkkxc'''''cK
                    O;'''',xWMMMMMMMMMXl'''''dNMMMMMWx,'''',kWMMMMMMMMMMM0:''''',oNMMMMMWk,''''dNKo,'''':O0c''''cKMMMMMMMXl'''''''''''''''cK
                    O;'''',xWMMMMMMMMMXl'''''dNMMMMMWx,'''',kWMMMMMMMMMMMKc''''''oNMMMMMWk,''',d0l'''''c0W0:''''cKMMMMMMMXo,,,,,,,,,''''''cK
                    Kc'''''lXMMMMMMMMMXl'''''dNMMMMMMO;'''''dNMMMMMMMMMMMO;''''',xWMMMMMWk,''',cc'''',oKMM0:''''cKMMMMMMMWX000000000o'''''cK
                    Wx,'''',dXMMMMMMMMXl'''''dNMMMMMMXl'''.';OWMMMMMMMMMKl''''':o0MMMMMMWk,''''''''',dNMMM0:''''cKMMMMMMMMMMMMMMMMMWx,''''cK
                    MNd,'''',ckXWMMMMMXl'''''dNMMMMMMM0c''''':xXWMMMMMNOc''''';kNWMMMMMMWk,'''''''':kNMMMM0:''''cKMMMMMMMMMMMMMMMMMWx,''''cK
                    MMNkc'''''';codxxxd;'''''dNMMMMMMMWKl,'''',:dkkOkdc,''''':OWMMMMMMMMWk,'''''''c0WMMMMM0:''''cKMMMMMMXkoddddddddo:'''''cK
                    MMMWXkl;,''''''''''''''''dNMMMMMMMMMNkl,''''''''''''''':dKWMMMMMMMMMWk,''''',lKWMMMMMM0:''''cKMMMMMM0:''''''''''''''''cK
                    MMMMMMWX0xdlc:;;;;,,;;;;:kWMMMMMMMMMMMN0xoc;;,,,,,;:ldOXWMMMMMMMMMMMMOc:::::dXMMMMMMMMKo::::dXMMMMMMKo::::::::::::::::oX
                    MMMMMMMMMMMWNXK000OO000KXNMMMMMMMMMMMMMMMWNKOkkxkO0XNMMMMMMMMMMMMMMMMWXKKKKXWMMMMMMMMMWXKKKKNWMMMMMMWXKKKKKKKKKKKKKKKKXW""")
        except:
            print("Something went wrong. Check the name of the file and make sure that the excel the second excel file is closed")
